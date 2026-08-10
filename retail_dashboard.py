"""
Retail Sales Analytics Dashboard
=================================
A command-line data pipeline for retail businesses that:
  1. Imports raw CSV sales data
  2. Cleans the data (duplicates, missing values, inconsistent formatting, invalid rows)
  3. Detects and reports missing values
  4. Performs sales analysis (revenue trends, regional performance, payment modes)
  5. Performs product performance analysis (top/bottom products, category performance)
  6. Generates charts (matplotlib, saved as PNG for embedding in reports)
  7. Exports a cleaned, multi-sheet Excel report
  8. Exports a formatted PDF summary report

Usage:
    python retail_dashboard.py <input_csv> [--outdir OUTPUT_DIR]

Example:
    python retail_dashboard.py sales_data_raw.csv --outdir output
"""

import argparse
import os
import sys
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
)


# --------------------------------------------------------------------------
# 1. DATA IMPORT
# --------------------------------------------------------------------------
def import_csv(path: str) -> pd.DataFrame:
    """Load a raw CSV file into a DataFrame."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Input CSV not found: {path}")
    df = pd.read_csv(path)
    print(f"[IMPORT] Loaded {len(df)} rows, {len(df.columns)} columns from {path}")
    return df


# --------------------------------------------------------------------------
# 2 & 3. MISSING VALUE DETECTION + DATA CLEANING
# --------------------------------------------------------------------------
def detect_missing_values(df: pd.DataFrame) -> pd.DataFrame:
    """Return a summary table of missing values per column."""
    missing_count = df.isna().sum()
    missing_pct = (missing_count / len(df) * 100).round(2)
    summary = pd.DataFrame({
        "Column": missing_count.index,
        "MissingCount": missing_count.values,
        "MissingPercent": missing_pct.values
    })
    summary = summary[summary["MissingCount"] > 0].sort_values(
        "MissingCount", ascending=False
    ).reset_index(drop=True)
    print("[MISSING VALUES]")
    if summary.empty:
        print("  No missing values found.")
    else:
        print(summary.to_string(index=False))
    return summary


def clean_data(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Clean the raw sales data:
      - Standardize column types
      - Trim/normalize text fields (Region, PaymentMode, Category, Product)
      - Drop exact duplicate rows
      - Drop rows with invalid/non-positive Quantity or UnitPrice
      - Impute missing numeric values (median) and categorical values (mode)
      - Add a computed Revenue column
    Returns the cleaned DataFrame and a dict of cleaning stats for reporting.
    """
    stats = {}
    df = df.copy()
    stats["rows_before"] = len(df)

    # --- Parse dates ---
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # --- Normalize text fields: trim whitespace, title case ---
    text_cols = ["Product", "Category", "Region", "PaymentMode", "CustomerID"]
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df.loc[df[col].isin(["nan", "None", ""]), col] = np.nan
            if col in ("Region", "Category", "PaymentMode"):
                df[col] = df[col].str.title()

    # --- Remove exact duplicate rows ---
    dup_count = df.duplicated().sum()
    df = df.drop_duplicates().reset_index(drop=True)
    stats["duplicates_removed"] = int(dup_count)

    # --- Remove rows with invalid Quantity (<= 0) ---
    invalid_qty = df["Quantity"].le(0).sum() if "Quantity" in df.columns else 0
    df = df[~(df["Quantity"].le(0))]
    stats["invalid_quantity_removed"] = int(invalid_qty)

    # --- Remove rows with missing/invalid Date or Product (can't analyze without these) ---
    missing_critical = df["Date"].isna().sum() + df["Product"].isna().sum()
    df = df.dropna(subset=["Date", "Product"])
    stats["missing_critical_removed"] = int(missing_critical)

    # --- Impute remaining missing numeric fields with median ---
    for col in ["Quantity", "UnitPrice"]:
        if col in df.columns and df[col].isna().any():
            median_val = df[col].median()
            n_filled = df[col].isna().sum()
            df[col] = df[col].fillna(median_val)
            stats[f"{col}_imputed"] = int(n_filled)

    # --- Impute remaining missing categorical fields with mode ---
    for col in ["Region", "PaymentMode", "Category"]:
        if col in df.columns and df[col].isna().any():
            mode_val = df[col].mode(dropna=True)
            fill_val = mode_val.iloc[0] if not mode_val.empty else "Unknown"
            n_filled = df[col].isna().sum()
            df[col] = df[col].fillna(fill_val)
            stats[f"{col}_imputed"] = int(n_filled)

    # --- CustomerID: fill remaining blanks with "Guest" ---
    if "CustomerID" in df.columns and df["CustomerID"].isna().any():
        n_filled = df["CustomerID"].isna().sum()
        df["CustomerID"] = df["CustomerID"].fillna("Guest")
        stats["CustomerID_imputed"] = int(n_filled)

    # --- Computed field ---
    df["Revenue"] = (df["Quantity"] * df["UnitPrice"]).round(2)
    df["Month"] = df["Date"].dt.to_period("M").astype(str)

    df = df.sort_values("Date").reset_index(drop=True)
    stats["rows_after"] = len(df)
    stats["rows_removed_total"] = stats["rows_before"] - stats["rows_after"]

    print("[CLEANING SUMMARY]")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    return df, stats


# --------------------------------------------------------------------------
# 4. SALES ANALYSIS
# --------------------------------------------------------------------------
def analyze_sales(df: pd.DataFrame) -> dict:
    """Compute overall sales KPIs and trends."""
    total_revenue = df["Revenue"].sum()
    total_orders = df["OrderID"].nunique()
    total_units = df["Quantity"].sum()
    avg_order_value = df.groupby("OrderID")["Revenue"].sum().mean()
    date_range = (df["Date"].min(), df["Date"].max())

    monthly_sales = df.groupby("Month").agg(
        Revenue=("Revenue", "sum"),
        Orders=("OrderID", "nunique"),
        UnitsSold=("Quantity", "sum")
    ).reset_index()

    region_sales = df.groupby("Region").agg(
        Revenue=("Revenue", "sum"),
        Orders=("OrderID", "nunique")
    ).reset_index().sort_values("Revenue", ascending=False)

    payment_sales = df.groupby("PaymentMode").agg(
        Revenue=("Revenue", "sum"),
        Orders=("OrderID", "nunique")
    ).reset_index().sort_values("Revenue", ascending=False)

    results = {
        "total_revenue": round(total_revenue, 2),
        "total_orders": int(total_orders),
        "total_units": int(total_units),
        "avg_order_value": round(avg_order_value, 2),
        "date_range": date_range,
        "monthly_sales": monthly_sales,
        "region_sales": region_sales,
        "payment_sales": payment_sales,
    }

    print("[SALES ANALYSIS]")
    print(f"  Total Revenue: ₹{total_revenue:,.2f}")
    print(f"  Total Orders: {total_orders}")
    print(f"  Total Units Sold: {total_units}")
    print(f"  Avg Order Value: ₹{avg_order_value:,.2f}")

    return results


# --------------------------------------------------------------------------
# 5. PRODUCT PERFORMANCE ANALYSIS
# --------------------------------------------------------------------------
def analyze_products(df: pd.DataFrame, top_n: int = 5) -> dict:
    """Rank products and categories by revenue and units sold."""
    product_perf = df.groupby("Product").agg(
        Revenue=("Revenue", "sum"),
        UnitsSold=("Quantity", "sum"),
        Orders=("OrderID", "nunique")
    ).reset_index().sort_values("Revenue", ascending=False).reset_index(drop=True)

    category_perf = df.groupby("Category").agg(
        Revenue=("Revenue", "sum"),
        UnitsSold=("Quantity", "sum"),
        Orders=("OrderID", "nunique")
    ).reset_index().sort_values("Revenue", ascending=False).reset_index(drop=True)

    top_products = product_perf.head(top_n)
    bottom_products = product_perf.tail(top_n).sort_values("Revenue")

    print("[PRODUCT PERFORMANCE]")
    print(f"  Top {top_n} products by revenue:")
    print(top_products.to_string(index=False))

    return {
        "product_perf": product_perf,
        "category_perf": category_perf,
        "top_products": top_products,
        "bottom_products": bottom_products,
    }


# --------------------------------------------------------------------------
# 6. CHARTS
# --------------------------------------------------------------------------
def generate_charts(sales_results: dict, product_results: dict, outdir: str) -> dict:
    """Generate PNG charts and return their file paths."""
    os.makedirs(outdir, exist_ok=True)
    chart_paths = {}
    plt.rcParams.update({"font.size": 10, "axes.titleweight": "bold"})

    # --- Monthly revenue trend (line chart) ---
    monthly = sales_results["monthly_sales"]
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.plot(monthly["Month"], monthly["Revenue"], marker="o", color="#2563eb", linewidth=2)
    ax.set_title("Monthly Revenue Trend")
    ax.set_xlabel("Month")
    ax.set_ylabel("Revenue (₹)")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x/1000:.0f}K"))
    plt.xticks(rotation=45, ha="right")
    fig.tight_layout()
    path = os.path.join(outdir, "chart_monthly_trend.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    chart_paths["monthly_trend"] = path

    # --- Top products (bar chart) ---
    top = product_results["top_products"]
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.barh(top["Product"][::-1], top["Revenue"][::-1], color="#16a34a")
    ax.set_title("Top Products by Revenue")
    ax.set_xlabel("Revenue (₹)")
    fig.tight_layout()
    path = os.path.join(outdir, "chart_top_products.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    chart_paths["top_products"] = path

    # --- Revenue by region (pie chart) ---
    region = sales_results["region_sales"]
    fig, ax = plt.subplots(figsize=(5.5, 4.5))
    ax.pie(region["Revenue"], labels=region["Region"], autopct="%1.1f%%",
           colors=["#2563eb", "#16a34a", "#f59e0b", "#dc2626", "#7c3aed"])
    ax.set_title("Revenue Share by Region")
    fig.tight_layout()
    path = os.path.join(outdir, "chart_region_share.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    chart_paths["region_share"] = path

    # --- Category performance (bar chart) ---
    cat = product_results["category_perf"]
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.bar(cat["Category"], cat["Revenue"], color="#f59e0b")
    ax.set_title("Revenue by Category")
    ax.set_ylabel("Revenue (₹)")
    plt.xticks(rotation=30, ha="right")
    fig.tight_layout()
    path = os.path.join(outdir, "chart_category.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    chart_paths["category"] = path

    print(f"[CHARTS] Saved {len(chart_paths)} charts to {outdir}")
    return chart_paths


# --------------------------------------------------------------------------
# 7. EXCEL EXPORT
# --------------------------------------------------------------------------
def export_excel(df, missing_summary, cleaning_stats, sales_results,
                  product_results, out_path: str):
    """Export cleaned data + analysis to a multi-sheet, formatted Excel workbook."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    title_font = Font(bold=True, size=14, name="Arial", color="1E3A8A")
    thin_border = Border(*(Side(style="thin", color="D1D5DB"),) * 4)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def autofit(ws, ncols, min_w=10, max_w=40):
        for c in range(1, ncols + 1):
            col_letter = get_column_letter(c)
            max_len = min_w
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)

    def write_df(ws, df_, start_row=1, title=None):
        r = start_row
        if title:
            ws.cell(row=r, column=1, value=title).font = title_font
            r += 2
        for j, col in enumerate(df_.columns, start=1):
            ws.cell(row=r, column=j, value=col)
        style_header(ws, r, len(df_.columns))
        for i, row_data in enumerate(df_.itertuples(index=False), start=1):
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r + i, column=j, value=val)
                cell.border = thin_border
        autofit(ws, len(df_.columns))
        return r + len(df_) + 1

    # --- Sheet 1: Summary / KPIs ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Retail Sales Dashboard — Summary Report"
    ws["A1"].font = Font(bold=True, size=16, color="1E3A8A")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    kpis = [
        ("Total Revenue (₹)", sales_results["total_revenue"]),
        ("Total Orders", sales_results["total_orders"]),
        ("Total Units Sold", sales_results["total_units"]),
        ("Average Order Value (₹)", sales_results["avg_order_value"]),
        ("Data Period Start", sales_results["date_range"][0].strftime("%Y-%m-%d")),
        ("Data Period End", sales_results["date_range"][1].strftime("%Y-%m-%d")),
        ("Rows Before Cleaning", cleaning_stats["rows_before"]),
        ("Rows After Cleaning", cleaning_stats["rows_after"]),
        ("Duplicates Removed", cleaning_stats["duplicates_removed"]),
        ("Rows Removed (invalid/critical)", cleaning_stats["invalid_quantity_removed"]
            + cleaning_stats["missing_critical_removed"]),
    ]
    r = 4
    ws.cell(row=r, column=1, value="Metric").font = header_font
    ws.cell(row=r, column=2, value="Value").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=2).fill = header_fill
    for i, (k, v) in enumerate(kpis, start=1):
        ws.cell(row=r + i, column=1, value=k).border = thin_border
        ws.cell(row=r + i, column=2, value=v).border = thin_border
    autofit(ws, 2)

    # --- Sheet 2: Cleaned Data ---
    ws2 = wb.create_sheet("Cleaned Data")
    export_cols = ["OrderID", "Date", "Product", "Category", "Quantity",
                   "UnitPrice", "Revenue", "Region", "PaymentMode", "CustomerID"]
    df_export = df[export_cols].copy()
    df_export["Date"] = df_export["Date"].dt.strftime("%Y-%m-%d")
    write_df(ws2, df_export, start_row=1)
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Missing Value Report ---
    ws3 = wb.create_sheet("Missing Values")
    if missing_summary.empty:
        ws3["A1"] = "No missing values detected in the raw dataset."
        ws3["A1"].font = Font(italic=True)
    else:
        write_df(ws3, missing_summary, start_row=1, title="Missing Value Report (raw data)")

    # --- Sheet 4: Sales Analysis ---
    ws4 = wb.create_sheet("Sales Analysis")
    r_end = write_df(ws4, sales_results["monthly_sales"], start_row=1, title="Monthly Sales")
    r_end = write_df(ws4, sales_results["region_sales"], start_row=r_end + 1, title="Sales by Region")
    write_df(ws4, sales_results["payment_sales"], start_row=r_end + 1, title="Sales by Payment Mode")

    # Add a native Excel bar chart for monthly revenue
    chart = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    n = len(sales_results["monthly_sales"])
    data = Reference(ws4, min_col=2, min_row=3, max_row=3 + n)
    cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 18, 9
    ws4.add_chart(chart, f"F3")

    # --- Sheet 5: Product Performance ---
    ws5 = wb.create_sheet("Product Performance")
    r_end = write_df(ws5, product_results["product_perf"], start_row=1, title="All Products — Ranked by Revenue")
    write_df(ws5, product_results["category_perf"], start_row=r_end + 1, title="Category Performance")

    bar = BarChart()
    bar.title = "Top Products by Revenue"
    bar.y_axis.title = "Revenue"
    top_n = min(10, len(product_results["product_perf"]))
    data = Reference(ws5, min_col=2, min_row=3, max_row=3 + top_n)
    cats = Reference(ws5, min_col=1, min_row=4, max_row=3 + top_n)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width, bar.height = 18, 9
    ws5.add_chart(bar, "F3")

    wb.save(out_path)
    print(f"[EXCEL EXPORT] Saved: {out_path}")


# --------------------------------------------------------------------------
# 8. PDF REPORT
# --------------------------------------------------------------------------
def export_pdf(sales_results, product_results, cleaning_stats, chart_paths, out_path: str):
    """Generate a formatted PDF business report with KPIs, tables, and charts."""
    doc = SimpleDocTemplate(out_path, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Title"], textColor=colors.HexColor("#1E3A8A"))
    h2 = ParagraphStyle("H2X", parent=styles["Heading2"], textColor=colors.HexColor("#2563EB"),
                         spaceBefore=14, spaceAfter=6)
    normal = styles["Normal"]
    story = []

    # --- Title page ---
    story.append(Paragraph("Retail Sales Analytics Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d %B %Y')}", normal))
    story.append(Spacer(1, 16))

    # --- KPI table ---
    story.append(Paragraph("Executive Summary", h2))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"₹{sales_results['total_revenue']:,.2f}"],
        ["Total Orders", f"{sales_results['total_orders']:,}"],
        ["Total Units Sold", f"{sales_results['total_units']:,}"],
        ["Average Order Value", f"₹{sales_results['avg_order_value']:,.2f}"],
        ["Reporting Period", f"{sales_results['date_range'][0].strftime('%d %b %Y')} – "
                              f"{sales_results['date_range'][1].strftime('%d %b %Y')}"],
    ]
    t = Table(kpi_data, colWidths=[7 * cm, 7 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    # --- Data quality note ---
    story.append(Paragraph("Data Cleaning Summary", h2))
    clean_text = (
        f"Of {cleaning_stats['rows_before']:,} raw records, "
        f"{cleaning_stats['duplicates_removed']} duplicate rows, "
        f"{cleaning_stats['invalid_quantity_removed']} rows with invalid quantities, and "
        f"{cleaning_stats['missing_critical_removed']} rows missing critical fields (date/product) "
        f"were removed. Remaining missing numeric and categorical values were imputed using "
        f"median and mode respectively, resulting in {cleaning_stats['rows_after']:,} clean records "
        f"used for analysis."
    )
    story.append(Paragraph(clean_text, normal))

    # --- Charts ---
    story.append(Paragraph("Monthly Revenue Trend", h2))
    story.append(Image(chart_paths["monthly_trend"], width=16 * cm, height=8 * cm))

    story.append(Paragraph("Revenue Share by Region", h2))
    story.append(Image(chart_paths["region_share"], width=12 * cm, height=9.5 * cm))

    story.append(PageBreak())

    story.append(Paragraph("Top Performing Products", h2))
    story.append(Image(chart_paths["top_products"], width=16 * cm, height=8 * cm))

    top_prod_data = [["Product", "Revenue (₹)", "Units Sold", "Orders"]]
    for _, row in product_results["top_products"].iterrows():
        top_prod_data.append([
            row["Product"], f"{row['Revenue']:,.2f}", int(row["UnitsSold"]), int(row["Orders"])
        ])
    t2 = Table(top_prod_data, colWidths=[6 * cm, 4 * cm, 3 * cm, 3 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16A34A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(Spacer(1, 8))
    story.append(t2)

    story.append(Paragraph("Revenue by Category", h2))
    story.append(Image(chart_paths["category"], width=16 * cm, height=8 * cm))

    # --- Insights ---
    story.append(Paragraph("Key Business Insights", h2))
    top_region = sales_results["region_sales"].iloc[0]
    top_product = product_results["top_products"].iloc[0]
    top_category = product_results["category_perf"].iloc[0]
    insights = [
        f"• {top_region['Region']} is the highest-performing region, generating "
        f"₹{top_region['Revenue']:,.2f} in revenue.",
        f"• \"{top_product['Product']}\" is the best-selling product by revenue "
        f"(₹{top_product['Revenue']:,.2f} from {int(top_product['UnitsSold'])} units).",
        f"• The \"{top_category['Category']}\" category leads all categories with "
        f"₹{top_category['Revenue']:,.2f} in revenue.",
        f"• Average order value stands at ₹{sales_results['avg_order_value']:,.2f}, "
        f"a useful benchmark for upsell and bundling strategies.",
    ]
    for line in insights:
        story.append(Paragraph(line, normal))
        story.append(Spacer(1, 4))

    doc.build(story)
    print(f"[PDF EXPORT] Saved: {out_path}")


# --------------------------------------------------------------------------
# MAIN PIPELINE
# --------------------------------------------------------------------------
def run_pipeline(input_csv: str, outdir: str = "output"):
    os.makedirs(outdir, exist_ok=True)

    raw_df = import_csv(input_csv)
    missing_summary = detect_missing_values(raw_df)
    clean_df, cleaning_stats = clean_data(raw_df)

    sales_results = analyze_sales(clean_df)
    product_results = analyze_products(clean_df)

    chart_paths = generate_charts(sales_results, product_results, outdir)

    clean_df.to_csv(os.path.join(outdir, "sales_data_cleaned.csv"), index=False)

    export_excel(
        clean_df, missing_summary, cleaning_stats, sales_results, product_results,
        os.path.join(outdir, "Sales_Report.xlsx")
    )
    export_pdf(
        sales_results, product_results, cleaning_stats, chart_paths,
        os.path.join(outdir, "Sales_Report.pdf")
    )

    print("\n[PIPELINE COMPLETE]")
    print(f"  Cleaned CSV : {outdir}/sales_data_cleaned.csv")
    print(f"  Excel Report: {outdir}/Sales_Report.xlsx")
    print(f"  PDF Report  : {outdir}/Sales_Report.pdf")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Retail Sales Analytics Dashboard")
    parser.add_argument("input_csv", help="Path to the raw sales CSV file")
    parser.add_argument("--outdir", default="output", help="Output directory")
    args = parser.parse_args()
    run_pipeline(args.input_csv, args.outdir)
